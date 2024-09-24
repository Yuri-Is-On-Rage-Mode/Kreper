#############################
# By: M. Hamza Sufyan (v-1) #
#############################

import requests
from bs4 import BeautifulSoup
from rich import print
from rich.console import Console
import argparse
import os
import re
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

console = Console()

class Kreper:
    def __init__(self, url, user_agent=None, ignore_robots=False):
        self.url = url
        self.visited = set()
        self.data = []
        self.headers = {'User-Agent': user_agent} if user_agent else {}
        self.ignore_robots = ignore_robots

    def simple_scrape(self):
        try:
            response = requests.get(self.url, headers=self.headers)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except requests.RequestException as e:
            console.print(f"[bold red]Error:[/bold red] {e}")
            return None

    def crawl(self, depth, limit):
        if depth < 0 or len(self.visited) >= limit or self.url in self.visited:
            return
        self.visited.add(self.url)
        console.print(f"[bold green]Crawling:[/bold green] {self.url}")

        soup = self.simple_scrape()
        if soup:
            for link in soup.find_all('a', href=True):
                next_url = link['href']
                if next_url.startswith('/'):
                    next_url = f"{self.url}{next_url}"
                if next_url not in self.visited:
                    self.crawl(depth - 1, limit)

    def extract_tags(self, tags):
        soup = self.simple_scrape()
        if soup:
            for tag in tags:
                elements = soup.find_all(tag)
                for element in elements:
                    console.print(f"[bold blue]{tag}:[/bold blue] {element.text.strip()}")
                    self.data.append({tag: element.text.strip()})

    def extract_images(self):
        soup = self.simple_scrape()
        if soup:
            images = soup.find_all('img')
            for img in images:
                img_src = img.get('src')
                if img_src:
                    console.print(f"[bold yellow]Image URL:[/bold yellow] {img_src}")
                    self.data.append({"image": img_src})

    def extract_video(self):
        soup = self.simple_scrape()
        if soup:
            videos = soup.find_all('video')
            for video in videos:
                video_src = video.get('src')
                if video_src:
                    console.print(f"[bold magenta]Video URL:[/bold magenta] {video_src}")
                    self.data.append({"video": video_src})

    def extract_audio(self):
        soup = self.simple_scrape()
        if soup:
            audios = soup.find_all('audio')
            for audio in audios:
                audio_src = audio.get('src')
                if audio_src:
                    console.print(f"[bold cyan]Audio URL:[/bold cyan] {audio_src}")
                    self.data.append({"audio": audio_src})

    def extract_meta(self):
        soup = self.simple_scrape()
        if soup:
            meta_tags = soup.find_all('meta')
            for meta in meta_tags:
                content = meta.get('content')
                name = meta.get('name')
                console.print(f"[bold green]Meta Tag:[/bold green] {name} - {content}")
                self.data.append({"meta": {name: content}})

    def extract_table(self):
        soup = self.simple_scrape()
        if soup:
            tables = soup.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                if rows:
                    # Extract table headers
                    headers = [th.text.strip() for th in rows[0].find_all('th')]
                    # Create a rich table with headers
                    rich_table = Table(title="Extracted Table")
                    for header in headers:
                        rich_table.add_column(header, justify="center")

                    # Extract table rows
                    for row in rows[1:]:
                        columns = [td.text.strip() for td in row.find_all('td')]
                        rich_table.add_row(*columns)
                    
                    console.print(rich_table)
                    self.data.append({"table": rich_table})

    def store_html(self):
        soup = self.simple_scrape()
        if soup:
            with open('page.html', 'w', encoding='utf-8') as f:
                f.write(str(soup))
            console.print("[bold yellow]HTML saved to:[/bold yellow] page.html")

    def search_text(self, search_term):
        soup = self.simple_scrape()
        if soup:
            if re.search(search_term, soup.text, re.IGNORECASE):
                console.print(f"[bold green]Found text:[/bold green] {search_term}")
            else:
                console.print(f"[bold red]Text not found:[/bold red] {search_term}")

    
    

    def output_excel(self, site_name: str, file_name: str, output_dir: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import os
        import hashlib

        """
        Outputs extracted data to an Excel file.

        Args:
        site_name (str): Name of the website.
        file_name (str): Name of the file.
        output_dir (str): Output directory.
        """

        # Create directory for extracted data
        directory = os.path.join(output_dir, site_name, '.extracted_data')
        os.makedirs(directory, exist_ok=True)

        # Generate hash for unique filename
        hash_object = hashlib.md5(file_name.encode())
        hash_filename = hash_object.hexdigest()

        # Create Excel file path
        file_path = os.path.join(directory, f"{hash_filename}.xlsx")

        # Initialize workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # Set header row font and alignment
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write header
        if self.data:
            headers = list(self.data[0].keys())
            for i, header in enumerate(headers):
                cell = ws.cell(row=1, column=i+1)
                cell.value = header
                cell.font = header_font
                cell.alignment = header_alignment

                # Set column width
                column_letter = get_column_letter(i+1)
                ws.column_dimensions[column_letter].width = 20
            
        # Write data
        for i, item in enumerate(self.data, start=2):
            for j, value in enumerate(item.values()):
                cell = ws.cell(row=i, column=j+1)
                cell.value = value
                
                # Set cell font and alignment
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Highlight alternate rows
                if i % 2 == 0:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        
        # Set border
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.rows:
            for cell in row:
                cell.border = border
        
        wb.save(file_path)
        console.print(f"[bold yellow]Data saved to:[/bold yellow] {file_path}")

    def download_images(self, output_dir):
        soup = self.simple_scrape()
        if soup:
            images = soup.find_all('img')
            for img in images:
                img_src = img.get('src')
                if img_src:
                    console.print(f"[bold yellow]Downloading image:[/bold yellow] {img_src}")
                    response = requests.get(img_src)
                    if response.status_code == 200:
                        filename = img_src.split('/')[-1]
                        with open(os.path.join(output_dir, filename), 'wb') as f:
                            f.write(response.content)
                    else:
                        console.print(f"[bold red]Failed to download image:[/bold red] {img_src}")

    def download_videos(self, output_dir):
        soup = self.simple_scrape()
        if soup:
            videos = soup.find_all('video')
            for video in videos:
                video_src = video.get('src')
                if video_src:
                    console.print(f"[bold magenta]Downloading video:[/bold magenta] {video_src}")
                    response = requests.get(video_src)
                    if response.status_code == 200:
                        filename = video_src.split('/')[-1]
                        with open(os.path.join(output_dir, filename), 'wb') as f:
                            f.write(response.content)
                    else:
                        console.print(f"[bold red]Failed to download video:[/bold red] {video_src}")

    def output_data(self, format: str, site_name: str, file_name: str, output_dir: str) -> None:
        import hashlib
        
        """
        Outputs extracted data to a file in the specified format.

        Args:
        format (str): Output format (json, csv, xlsx).
        site_name (str): Name of the website.
        file_name (str): Name of the file.
        output_dir (str): Output directory.
        """

        # Create directory for extracted data
        directory = os.path.join(output_dir, site_name, '.extracted_data')
        os.makedirs(directory, exist_ok=True)

        # Generate hash for unique filename
        hash_object = hashlib.md5(file_name.encode())
        hash_filename = hash_object.hexdigest()

        try:
            if format == 'json':
                # Output data to JSON file
                file_path = os.path.join(directory, f"{hash_filename}.json")
                with open(file_path, 'w') as f:
                    json.dump(self.data, f, indent=4)
                console.print(f"[bold yellow]Data saved to:[/bold yellow] {file_path}")

            elif format == 'xlsx':
                # Output data to Excel file
                self.output_excel(site_name, file_name, output_dir)

            elif format == 'csv':
                # Output data to CSV file
                file_path = os.path.join(directory, f"{hash_filename}.csv")
                with open(file_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    # Write header
                    if self.data:
                        writer.writerow(self.data[0].keys())
                    for item in self.data:
                        writer.writerow(item.values())
                console.print(f"[bold yellow]Data saved to:[/bold yellow] {file_path}")

        except Exception as e:
            console.print(f"[bold red]Error saving data:[/bold red] {e}")


def main():
    parser = argparse.ArgumentParser(description="Web Scraper - Kreper")
    parser.add_argument('-u', '--url', help='Target URL to scrape')
    parser.add_argument('--crawl', action='store_true', help='Crawl the website')
    parser.add_argument('--depth', type=int, default=0, help='Depth for crawling')
    parser.add_argument('--limit', type=int, default=100, help='Limit number of pages to crawl')
    parser.add_argument('--extract', nargs='+', help='Extract specific HTML tags')
    parser.add_argument('--extract-images', action='store_true', help='Extract image URLs')
    parser.add_argument('--extract-video', action='store_true', help='Extract video URLs')
    parser.add_argument('--extract-audio', action='store_true', help='Extract audio URLs')
    parser.add_argument('--extract-meta', action='store_true', help='Extract meta tags')
    parser.add_argument('--extract-table', action='store_true', help='Extract tables from the page')
    parser.add_argument('--download-images', action='store_true', help='Download found images')
    parser.add_argument('--download-videos', action='store_true', help='Download found videos')
    parser.add_argument('--media-dir', type=str, default='./media', help='Directory for saving downloaded media')
    parser.add_argument('--store-html', action='store_true', help='Store complete HTML of the page')
    parser.add_argument('--search', type=str, help='Search for specific text in the page content')
    parser.add_argument('--output', type=str, choices=['json', 'csv', 'xlsx'], help='Output format for extracted data')
    parser.add_argument('--S', action='store_true', help='Save extracted data to specified directory')
    parser.add_argument('--file-name', type=str, default='output', help='Filename for saved output')
    parser.add_argument('--user-agent', type=str, help='Custom user-agent string')
    parser.add_argument('--ignore-robots', action='store_true', help='Ignore robots.txt rules')
    parser.add_argument('--output-dir', type=str, default='./', help='Directory for saving extracted data')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose output')
    parser.add_argument('--mc', type=str, help='Load commands from a script file')

    args = parser.parse_args()

    # Create instance of Kreper, URL will be set later if using --mc
    scraper = Kreper(url=args.url, user_agent=args.user_agent, ignore_robots=args.ignore_robots)

    if args.mc:
        try:
            with open(args.mc, 'r') as f:
                for line in f:
                    # Skip comments and empty lines
                    if line.startswith('//') or not line.strip():
                        continue
                    
                    # Execute command
                    command = line.strip()
                    command_parts = command.split()

                    # Handle -u argument to set URL from script
                    if command_parts[0] == '-u':
                        scraper.url = command_parts[1]
                    elif '--crawl' in command_parts:
                        scraper.crawl(args.depth, args.limit)
                    elif '--extract' in command_parts:
                        tags = command_parts[command_parts.index('--extract') + 1].strip('[]').split(',')
                        scraper.extract_tags(tags)
                    elif '--extract-images' in command_parts:
                        scraper.extract_images()
                    elif '--extract-video' in command_parts:
                        scraper.extract_video()
                    elif '--extract-audio' in command_parts:
                        scraper.extract_audio()
                    elif '--extract-table' in command_parts:
                        scraper.extract_table()
                    elif '--extract-meta' in command_parts:
                        scraper.extract_meta()
                    elif '--store-html' in command_parts:
                        scraper.store_html()
                    elif '--search' in command_parts:
                        search_term = command_parts[command_parts.index('--search') + 1]
                        scraper.search_text(search_term)
                    elif '--output' in command_parts:
                        output_format = command_parts[command_parts.index('--output') + 1]
                        scraper.output_data(output_format, scraper.url.split("//")[-1].split("/")[0], args.file_name, args.output_dir)

        except Exception as e:
            console.print(f"[bold red]Error loading commands from file:[/bold red] {e}")
            return

    # If URL was provided in command line, use it
    if args.url:
        scraper.url = args.url

    # Run commands if specified in command line
    if args.crawl:
        scraper.crawl(args.depth, args.limit)

    if args.extract:
        scraper.extract_tags(args.extract)

    if args.extract_images:
        scraper.extract_images()

    if args.extract_video:
        scraper.extract_video()

    if args.extract_audio:
        scraper.extract_audio()

    if args.extract_meta:
        scraper.extract_meta()

    if args.store_html:
        scraper.store_html()

    if args.search:
        scraper.search_text(args.search)

    if args.output:
        scraper.output_data(args.output, scraper.url.split("//")[-1].split("/")[0], args.file_name, args.output_dir)

if __name__ == '__main__':
    main()
